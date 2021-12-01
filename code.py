from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
import undetected_chromedriver as uc
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import csv
import random
from openpyxl import Workbook
import chromedriver_autoinstaller
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.by import By
import os
import urllib
import math
import sys
import time
import requests
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import tkinter as tk
from tkinter import simpledialog

ROOT = tk.Tk()

ROOT.withdraw()
# the input dialog
USER_INP = simpledialog.askstring(title="parola chiave",
                                  prompt="inserisci la parola chiave che vuoi cercare:")

# check it out
print(USER_INP)
audioToTextDelay = 10
delayTime = 2
audioFile = "\\payload.mp3"
URL = "https://www.google.com/recaptcha/api2/demo"
SpeechToTextURL = "https://speech-to-text-demo.ng.bluemix.net/"

def delay():
    time.sleep(random.randint(2, 3))

def audioToText(audioFile):
    driver.execute_script('''window.open("","_blank")''')
    driver.switch_to.window(driver.window_handles[1])
    driver.get(SpeechToTextURL)

    delay()
    audioInput = driver.find_element(By.XPATH, '//*[@id="root"]/div/input')
    audioInput.send_keys(audioFile)

    time.sleep(audioToTextDelay)

    text = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[7]/div/div/div/span')
    while text is None:
        text = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[7]/div/div/div/span')

    result = text.text

    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    return result
def captch_solve():
    iframes = driver.find_elements_by_tag_name('iframe')
    audioBtnFound = False
    audioBtnIndex = -1

    for index in range(len(iframes)):
        driver.switch_to.default_content()
        iframe = driver.find_elements_by_tag_name('iframe')[index]
        driver.switch_to.frame(iframe)
        driver.implicitly_wait(delayTime)
        try:
            audioBtn = driver.find_element_by_id("recaptcha-audio-button")
            audioBtn.click()
            audioBtnFound = True
            audioBtnIndex = index
            break
        except Exception as e:
            pass

    if audioBtnFound:
        try:
            while True:
                # get the mp3 audio file
                src = driver.find_element_by_id("audio-source").get_attribute("src")
                print("[INFO] Audio src: %s" % src)

                # download the mp3 audio file from the source
                urllib.request.urlretrieve(src, os.getcwd() + audioFile)

                # Speech To Text Conversion
                key = audioToText(os.getcwd() + audioFile)
                print("[INFO] Recaptcha Key: %s" % key)

                driver.switch_to.default_content()
                iframe = driver.find_elements_by_tag_name('iframe')[audioBtnIndex]
                driver.switch_to.frame(iframe)

                # key in results and submit
                inputField = driver.find_element_by_id("audio-response")
                inputField.send_keys(key)
                delay()
                inputField.send_keys(Keys.ENTER)
                delay()

                err = driver.find_elements_by_class_name('rc-audiochallenge-error-message')[0]
                if err.text == "" or err.value_of_css_property('display') == 'none':
                    print("[INFO] Success!")
                    break

        except Exception as e:
            print(e)
            sys.exit("[INFO] Possibly blocked by google. Change IP,Use Proxy method for requests")
    else:
        sys.exit("[INFO] Audio Play Button not found! In Very rare cases!")


options = webdriver.ChromeOptions() 
options.add_argument("start-maximized")
chromedriver_autoinstaller.install()
driver = uc.Chrome(options=options)
driver.get('https://www.registroimprese.it/home')
driver.implicitly_wait(30)
time.sleep(2)
driver.find_element_by_xpath('//input[@value="Acconsento"]').click()
time.sleep(2)
driver.find_element_by_id('inputSearchField').send_keys(USER_INP+ Keys.ENTER)
time.sleep(5)
try_again=False
try:
    driver.switch_to.default_content()
    driver.find_element_by_xpath('//a[@title="20 Items per Page"]').click()
    driver.find_element_by_xpath('//a[@id="_ricercaportlet_WAR_ricercaRIportlet_tiym_column5_0_menu_75"]').click()
except:
    try_again=True
    try:
        captch_solve()
    except:
        pass
finally:
    if try_again==True:        
        driver.switch_to.default_content()
        driver.implicitly_wait(30)
        driver.find_element_by_xpath('//a[@title="20 Items per Page"]').click()
        driver.find_element_by_xpath('//a[@id="_ricercaportlet_WAR_ricercaRIportlet_tiym_column5_0_menu_75"]').click()
time.sleep(2)
all_data=[]
total_pages=int(math.floor(int(driver.find_element_by_class_name('risTot').text.replace('.','')))/75)

for i in range(total_pages):
    if driver.find_element_by_xpath("//a[contains(text(), 'Successivo')]").get_attribute('href')=='javascript:;':
            driver.find_element_by_xpath('//div[@class="g-recaptcha"]').click() 
            time.sleep(3)
            try:
                driver.find_element_by_xpath("//button[contains(text(), 'PROCEDI')]").click()
                driver.implicitly_wait(30)
            except:
                try:
                    captch_solve()
                except:
                    pass
                finally:
                    driver.switch_to.default_content()
                    time.sleep(5)
                    driver.find_element_by_xpath("//button[contains(text(), 'PROCEDI')]").click()
                    driver.implicitly_wait(30)
                    time.sleep(5)
                    print(driver.find_element_by_xpath("//a[contains(text(), 'Successivo')]").get_attribute('href'))
                    if driver.find_element_by_xpath("//a[contains(text(), 'Successivo')]").get_attribute('href')=='javascript:;':
                        driver.refresh()
                        driver.implicitly_wait(30)
                        driver.find_element_by_xpath('//div[@class="g-recaptcha"]').click()
                        time.sleep(3)
                        try:
                            driver.find_element_by_xpath("//button[contains(text(), 'PROCEDI')]").click()
                            driver.implicitly_wait(30)
                        except:
                            try:
                                captch_solve()
                            except:
                                pass
                            finally:
                                driver.switch_to.default_content()
                                driver.find_element_by_xpath("//button[contains(text(), 'PROCEDI')]").click()
                                driver.implicitly_wait(30)
                                driver.sleep(5)

    
    for indexer,i in enumerate(driver.find_elements_by_xpath('//table[@class="table tableRisultatiGratuita"]/tbody/tr')):
        current_data=driver.find_elements_by_xpath('//table[@class="table tableRisultatiGratuita"]/tbody/tr')[indexer].text.splitlines()
        if len(current_data)==7:
            current_data.remove('... Leggi tutto')
        if len(current_data)==5:
            current_data.insert(-2,'Nothing')
        driver.find_elements_by_xpath('//table[@class="table tableRisultatiGratuita"]/tbody/tr')[indexer].find_element_by_tag_name('img').click()
        temp_list=[]
        for j in driver.find_elements_by_tag_name('dd'):
              temp_list.append(j.text.replace('Mostra mappa','').strip()) 
        try:
            temp_list.remove('')
        except:
            pass
        if 'MOSTRA' in temp_list:
            current_data.insert(1,temp_list[1])
            current_data.append(temp_list[-3])
            current_data.append(temp_list[-2])
            current_data.append(temp_list[-1])
            current_data[5]=temp_list[4]
            email=(driver.find_element_by_xpath('//a[@class="linkRisultatiRicerca showModalPec"]').get_attribute('onclick')).split('(')[-1]
            email=email.replace(')','')
            email=email.replace("'",'').strip()
            current_data.append(email)
            print(current_data)
        elif 'Non presente' in temp_list:
            current_data.insert(1,temp_list[1])
            current_data.append(temp_list[-3])
            current_data.append(temp_list[-2])
            current_data.append(temp_list[-1])
            current_data[5]=temp_list[4]
            email='Non presente'
            current_data.append(email)
            print(current_data)
        else:
            current_data.insert(1,temp_list[1])
            current_data.append(temp_list[-3])
            current_data.append(temp_list[-2])
            current_data.append(temp_list[-1])
            current_data[5]=temp_list[3]
            email='Non presente'
            current_data.append(email)
            print(current_data)      
        print(len(current_data))
        if len(current_data)!=11:
            raise Exception('More items')
        all_data.append(current_data)
        driver.back()
        driver.implicitly_wait(30)
    print(len(all_data))
    print('Page no.',indexer)
    driver.execute_script("scrollBy(0,-1000000000000000);")
    driver.find_element_by_xpath("//a[contains(text(), 'Successivo')]").click()
print('Data extraction completed')
time.sleep(2)
bold_font = Font(bold=True)
bold_font = Font(bold=True)
center_aligned_text = Alignment(horizontal="center")
double_border_side = Side(border_style="double")
square_border = Border(top=double_border_side,
                right=double_border_side,
                bottom=double_border_side,
                left=double_border_side)
workbook = Workbook()
sheet = workbook.active
['Pompe Funebri Sarni', 'Vallata (AV) Corso Kennedy 70/A', 'Sede Legale', 'Avellino', 'Altre Forme', "Servizi di pompe funebri e attivita' connesse", 'Registrata', '96.03', '96.03', '-', 'sarnigerardo80@sicurezzapostale.it']
sheet["A1"] = "Nome impresa"
sheet["A1"].font = bold_font
sheet["A1"].alignment = center_aligned_text
sheet["A1"].border = square_border
sheet["B1"] = "Indirizzo Impresa"
sheet["B1"].font = bold_font
sheet["B1"].alignment = center_aligned_text
sheet["B1"].border = square_border
sheet["C1"] = "Stato dell'ufficio"
sheet["C1"].font = bold_font
sheet["C1"].alignment = center_aligned_text
sheet["C1"].border = square_border
sheet["D1"] = "Provincia"
sheet["D1"].font = bold_font
sheet["D1"].alignment = center_aligned_text
sheet["D1"].border = square_border
sheet["E1"] = "Forma giuridica (generico)"
sheet["E1"].font = bold_font
sheet["E1"].alignment = center_aligned_text
sheet["E1"].border = square_border
sheet["F1"] = "Attività"
sheet["F1"].font = bold_font
sheet["F1"].alignment = center_aligned_text
sheet["F1"].border = square_border
sheet["G1"] = "Registrato/Non registrato"
sheet["G1"].font = bold_font
sheet["G1"].alignment = center_aligned_text
sheet["G1"].border = square_border
sheet["H1"] = "ATECO prevalente"
sheet["H1"].font = bold_font
sheet["H1"].alignment = center_aligned_text
sheet["H1"].border = square_border
sheet["I1"] = "ATECO primaria"
sheet["I1"].font = bold_font
sheet["I1"].alignment = center_aligned_text
sheet["I1"].border = square_border
sheet["J1"] = "ATECO secondaria"
sheet["J1"].font = bold_font
sheet["J1"].alignment = center_aligned_text
sheet["J1"].border = square_border
sheet["K1"] = "DOMICILIO DIGITALE / PEC*"
sheet["K1"].font = bold_font
sheet["K1"].alignment = center_aligned_text
sheet["K1"].border = square_border

dim_holder = DimensionHolder(worksheet=sheet)

for col in range(sheet.min_column, sheet.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

sheet.column_dimensions = dim_holder
for p,q in enumerate(all_data):
    sheet[f"A{p+2}"]=q[0]           
    sheet[f"B{p+2}"]=q[1]
    sheet[f"C{p+2}"]=q[2]
    sheet[f"D{p+2}"]=q[3]
    sheet[f"E{p+2}"]=q[4]
    sheet[f"F{p+2}"]=q[5]
    sheet[f"G{p+2}"]=q[6]
    sheet[f"H{p+2}"]=q[7]
    sheet[f"I{p+2}"]=q[8]
    sheet[f"J{p+2}"]=q[9]
    sheet[f"K{p+2}"]=q[10]
    
workbook.save(f"registroimprese_[{USER_INP}].xlsx")
print('process completed')